from __future__ import annotations

import json
import logging
import os
import unicodedata
from functools import lru_cache
from pathlib import Path
from typing import Any, Dict, Iterable, Optional
from urllib.parse import urlencode

import requests
from flask import current_app

try:
    import reverse_geocoder as rg
except ImportError:  # pragma: no cover - fallback caso dep opcional não esteja instalada
    rg = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - fallback caso openpyxl não esteja disponível
    load_workbook = None


# -------------------------------
# Endpoints base (v3 Agregados)
# -------------------------------
IBGE_LOCATIONS_URL = "https://servicodados.ibge.gov.br/api/v1/localidades/municipios"
# variável 93 (População residente), Censo 2022
IBGE_DEMOGRAPHICS_BASE = (
    "https://servicodados.ibge.gov.br/api/v3/agregados/9514/periodos/2022/variaveis/93"
)
# Fallback "legacy" para população total
IBGE_POPULATION_URL = (
    "https://servicodados.ibge.gov.br/api/v3/agregados/6579/periodos/2022/variaveis/9324?localidades=N6[{code}]"
)

# -------------------------------
# Classificações (IDs conforme metadado da 9514)
# -------------------------------
# Sexo (id da classificação = 2)
SEX_CLASS_ID = 2
SEX_IDS = [4, 5]  # 4 = Homens, 5 = Mulheres

# Idade (id da classificação = 287)
AGE_CLASS_ID = 287
# Faixas "nível 1" (consolidadas) + "100 anos ou mais"
AGE_BAND_IDS = [
    93070,  # 0 a 4
    93084,  # 5 a 9
    93085,  # 10 a 14
    93086,  # 15 a 19
    93087,  # 20 a 24
    93088,  # 25 a 29
    93089,  # 30 a 34
    93090,  # 35 a 39
    93091,  # 40 a 44
    93092,  # 45 a 49
    93093,  # 50 a 54
    93094,  # 55 a 59
    93095,  # 60 a 64
    93096,  # 65 a 69
    93097,  # 70 a 74
    93098,  # 75 a 79
    49108,  # 80 a 84
    49109,  # 85 a 89
    60040,  # 90 a 94
    60041,  # 95 a 99
    6653,   # 100 anos ou mais
]

# Forma de declaração da idade (id = 286)
# Quando omitida, a API costuma retornar "Total". Mantemos sem filtrar.


STATE_ALIASES = {
    "acre": "AC",
    "alagoas": "AL",
    "amapa": "AP",
    "amazonas": "AM",
    "bahia": "BA",
    "ceara": "CE",
    "distrito federal": "DF",
    "espirito santo": "ES",
    "goias": "GO",
    "maranhao": "MA",
    "mato grosso": "MT",
    "mato grosso do sul": "MS",
    "minas gerais": "MG",
    "para": "PA",
    "paraiba": "PB",
    "parana": "PR",
    "pernambuco": "PE",
    "piaui": "PI",
    "rio de janeiro": "RJ",
    "rio grande do norte": "RN",
    "rio grande do sul": "RS",
    "rondonia": "RO",
    "roraima": "RR",
    "santa catarina": "SC",
    "sao paulo": "SP",
    "sergipe": "SE",
    "tocantins": "TO",
}


ROOT_DIR = Path(__file__).resolve().parents[2]
_LOCAL_DATASET_PATH = ROOT_DIR / "data" / "ibge_population_income.json"
_POPULATION_XLSX = ROOT_DIR / "docs" / "CD2022_Populacao_Coletada_Imputada_e_Total_Municipio_e_UF_20231222.xlsx"
_INCOME_XLSX = ROOT_DIR / "docs" / "Agregados_por_municipios_renda_responsavel_BR.xlsx"
_LOCAL_DATASET: Dict[str, Any] | None = None
_RG_CLIENT = None

ALLOW_REMOTE_LOOKUPS = os.environ.get("IBGE_ALLOW_REMOTE", "0").lower() in ("1", "true", "yes")


def _log(event: str, level: int = logging.WARNING, **extra):
    """Logger seguro (funciona dentro/fora do contexto Flask)."""
    logger = logging.getLogger("ibge_service")
    try:
        logger = current_app.logger  # type: ignore[assignment]
    except Exception:
        pass

    if level <= logging.DEBUG:
        logger.debug(event, extra=extra)
    elif level == logging.INFO:
        logger.info(event, extra=extra)
    elif level >= logging.ERROR:
        logger.error(event, extra=extra)
    else:
        logger.warning(event, extra=extra)


def _slugify(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value or "")
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii").lower()
    for old, new in (("state of ", ""), ("estado de ", "")):
        ascii_text = ascii_text.replace(old, new)
    ascii_text = ascii_text.replace("-", " ").replace("_", " ")
    return " ".join(ascii_text.split())


def _coerce_int(value) -> int | None:
    if value in (None, "", "-", "..."):
        return None
    try:
        if isinstance(value, (int, float)):
            return int(value)
        cleaned = str(value).strip().replace(".", "").replace(",", ".")
        return int(float(cleaned))
    except (TypeError, ValueError):
        return None


def _coerce_float(value) -> float | None:
    if value in (None, "", "-", "..."):
        return None
    try:
        if isinstance(value, (int, float)):
            return float(value)
        cleaned = str(value).strip().replace(" ", "").replace(".", "").replace(",", ".")
        return float(cleaned)
    except (TypeError, ValueError):
        return None


def _build_dataset_from_xlsx() -> Dict[str, Any]:
    dataset: Dict[str, Any] = {}
    if load_workbook is None:
        return dataset

    if _POPULATION_XLSX.exists():
        wb = load_workbook(_POPULATION_XLSX, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=3, values_only=True):
            uf = row[1]
            cod_uf = row[2]
            cod_munic = row[3]
            name = row[4]
            total = row[7] if len(row) > 7 else None
            if not uf or uf == "UF":
                continue
            try:
                uf_code = int(cod_uf)
                city_code = str(cod_munic).zfill(5)
                code = f"{uf_code:02d}{city_code}"
            except (TypeError, ValueError):
                continue
            population = _coerce_int(total)
            if population is None:
                continue
            entry_code = str(code)
            dataset[entry_code] = {
                "code": entry_code,
                "name": name,
                "state": uf,
                "state_code": f"{uf_code:02d}",
                "population": population,
                "population_year": 2022,
            }

    if _INCOME_XLSX.exists():
        wb = load_workbook(_INCOME_XLSX, read_only=True, data_only=True)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = row[0]
            if not code:
                continue
            entry = dataset.setdefault(str(code), {"code": str(code)})
            per_capita = _coerce_float(row[5] if len(row) > 5 else None)
            total_income = _coerce_float(row[6] if len(row) > 6 else None)
            if per_capita is not None:
                entry["income_per_capita"] = per_capita
                entry["income_year"] = 2022
            if total_income is not None:
                entry["income_total"] = total_income
                entry["income_total_year"] = 2022

    if dataset:
        payload = {
            "generated_at": None,
            "municipalities": dataset,
        }
        try:
            _LOCAL_DATASET_PATH.parent.mkdir(parents=True, exist_ok=True)
            _LOCAL_DATASET_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2))
        except OSError:
            pass
    return dataset


def _load_local_dataset() -> Dict[str, Any]:
    global _LOCAL_DATASET
    if _LOCAL_DATASET is not None:
        return _LOCAL_DATASET
    if not _LOCAL_DATASET_PATH.exists():
        if _POPULATION_XLSX.exists() and _INCOME_XLSX.exists():
            _LOCAL_DATASET = _build_dataset_from_xlsx()
        else:
            _LOCAL_DATASET = {}
        return _LOCAL_DATASET
    try:
        payload = json.loads(_LOCAL_DATASET_PATH.read_text())
        municipalities = payload.get("municipalities") or {}
        normalized: Dict[str, Any] = {}
        for code, data in municipalities.items():
            if isinstance(data, dict):
                data.setdefault("code", str(code))
                if data.get("code"):
                    data.setdefault("state_code", str(data["code"])[0:2])
            normalized[str(code)] = data
        _LOCAL_DATASET = normalized
    except Exception as exc:
        _log("ibge.local_dataset.load_failed", level=logging.ERROR, error=str(exc))
        if _POPULATION_XLSX.exists() and _INCOME_XLSX.exists():
            _LOCAL_DATASET = _build_dataset_from_xlsx()
        else:
            _LOCAL_DATASET = {}
    return _LOCAL_DATASET


def _local_lookup_by_code(code: str | None) -> Dict[str, Any] | None:
    if not code:
        return None
    dataset = _load_local_dataset()
    entry = dataset.get(str(code))
    if entry and isinstance(entry, dict):
        entry.setdefault("code", str(code))
        if entry.get("code"):
            entry.setdefault("state_code", str(entry["code"])[0:2])
    return entry


def _local_lookup_by_city(city: str | None, state: str | None) -> Dict[str, Any] | None:
    if not city:
        return None
    dataset = _load_local_dataset()
    if not dataset:
        return None
    normalized_state = normalize_state_code(state) if state else None
    city_slug = _slugify(city)
    for entry in dataset.values():
        if _slugify(entry.get("name")) != city_slug:
            continue
        if normalized_state and normalize_state_code(entry.get("state")) != normalized_state:
            continue
        entry.setdefault("code", entry.get("code"))
        if entry.get("code"):
            entry.setdefault("state_code", str(entry["code"])[0:2])
        return entry
    return None


class ReverseGeocoderUnavailable(RuntimeError):
    """Erro lançado quando a dependência de geocodificação offline não está disponível."""


def _get_reverse_geocoder():
    global _RG_CLIENT
    if _RG_CLIENT is not None:
        return _RG_CLIENT
    if rg is None:
        raise ReverseGeocoderUnavailable(
            "Pacote reverse_geocoder não instalado. Execute pip install reverse_geocoder."
        )
    try:
        # modo single-process evita /dev/shm; modo 1 usa shared memory quando disponível
        try:
            _RG_CLIENT = rg.RGeocoder(mode=1, verbose=False)
        except Exception:
            _RG_CLIENT = rg.RGeocoder(mode=0, verbose=False)
    except Exception as exc:
        raise ReverseGeocoderUnavailable(f"Reverse geocoder init failed: {exc}") from exc
    return _RG_CLIENT


def reverse_geocode_offline(lat: float, lon: float) -> Dict[str, Any]:
    geocoder = _get_reverse_geocoder()
    try:
        result = geocoder.query([(float(lat), float(lon))])[0]
    except Exception as exc:  # pragma: no cover - erros raros do dataset
        raise ReverseGeocoderUnavailable(f"Reverse geocode falhou: {exc}") from exc
    state_label = result.get("adm1") or result.get("admin1") or result.get("adm2")
    state_code = normalize_state_code(state_label)
    return {
        "name": result.get("name"),
        "state": state_label,
        "state_code": state_code,
        "country": result.get("cc"),
        "provider": "reverse_geocoder",
    }


def get_local_municipality_entry(code: str | None) -> Dict[str, Any] | None:
    entry = _local_lookup_by_code(code)
    return entry


def find_local_municipality(city: str | None, state: str | None = None) -> Dict[str, Any] | None:
    return _local_lookup_by_city(city, state)


def normalize_state_code(value: str | None) -> str | None:
    if not value:
        return None
    v = value.strip()
    if len(v) == 2 and v.isalpha():
        return v.upper()
    slug = _slugify(v)
    return STATE_ALIASES.get(slug)


def _safe_parse_numeric(value) -> int | None:
    if value in (None, "", "-", "NA"):
        return None
    try:
        # a API retorna inteiros; isto garante robustez se vier "1234,0"
        return int(float(str(value).replace(",", ".")))
    except (TypeError, ValueError):
        return None


def _build_demographics_url(code: str, classificacoes: Optional[dict[int, Iterable[int]]] = None) -> str:
    """
    Monta a URL da variável 93 para um município N6[code], opcionalmente
    adicionando parâmetros de classificação, e.g.:
      classificacoes = { 2: [4,5], 287: [93087,93088,...] }
    vira:
      &classificacao=2[4,5]&classificacao=287[93087,93088,...]
    """
    base = IBGE_DEMOGRAPHICS_BASE
    params = {"localidades": f"N6[{code}]"}
    qs = "?" + urlencode(params)

    if classificacoes:
        # a API aceita repetir 'classificacao='
        parts = [f"classificacao={cid}[{','.join(str(x) for x in values)}]"
                 for cid, values in classificacoes.items()
                 if values]
        if parts:
            qs += "&" + "&".join(parts)

    return base + qs


@lru_cache(maxsize=2048)
def resolve_municipality_code(city: str | None, state: str | None = None) -> str | None:
    if not city:
        return None
    entry = _local_lookup_by_city(city, state)
    if entry and entry.get("code"):
        return str(entry["code"])
    if not ALLOW_REMOTE_LOOKUPS:
        return None
    # normaliza a consulta para melhor acerto/cache
    params = {"nome": _slugify(city)}
    try:
        resp = requests.get(IBGE_LOCATIONS_URL, params=params, timeout=10)
        resp.raise_for_status()
        candidates = resp.json()
    except Exception as exc:
        _log("ibge.lookup_failed", city=city, state=state, error=str(exc))
        entry = _local_lookup_by_city(city, state)
        return str(entry["code"]) if entry and entry.get("code") else None

    if not candidates:
        entry = _local_lookup_by_city(city, state)
        return str(entry["code"]) if entry and entry.get("code") else None

    normalized_state = normalize_state_code(state) if state else None

    def _candidate_state(item):
        uf_info = (((item.get("microrregiao") or {}).get("mesorregiao") or {}).get("UF") or {})
        sigla = uf_info.get("sigla")
        nome = uf_info.get("nome")
        return normalize_state_code(sigla) or normalize_state_code(nome)

    for item in candidates:
        candidate_state = _candidate_state(item)
        if not normalized_state or candidate_state == normalized_state:
            code = item.get("id")
            if code:
                return str(code)

    code = candidates[0].get("id")
    if code:
        return str(code)
    entry = _local_lookup_by_city(city, state)
    return str(entry["code"]) if entry and entry.get("code") else None


def _parse_total_from_payload(payload) -> Optional[int]:
    """
    Lê o 'total' (sem classificação) quando presente no payload padrão da 9514.
    """
    if not isinstance(payload, list) or not payload:
        return None
    resultados = payload[0].get("resultados", []) if isinstance(payload[0], dict) else []
    for resultado in resultados:
        series = resultado.get("series") or []
        for serie in series:
            classes = serie.get("classificacoes") or []
            # total típico: sem classificações
            if not classes:
                val = None
                for raw in (serie.get("serie") or {}).values():
                    val = _safe_parse_numeric(raw)
                    if val is not None:
                        return val
    return None


def _parse_breakdown(payload, target_class_name: str) -> Dict[str, int]:
    """
    Soma valores por rótulo de classe desejada (e.g. 'sexo' OU 'idade'),
    assumindo que a outra(s) dimensão(ões) não foram pedidas (para evitar dupla contagem).
    """
    out: Dict[str, int] = {}
    if not isinstance(payload, list) or not payload:
        return out

    resultados = payload[0].get("resultados", []) if isinstance(payload[0], dict) else []
    for resultado in resultados:
        # mapeia metadados das classificações deste bloco
        klasses = resultado.get("classificacoes") or []
        classes_meta: dict[str, dict[str, Any]] = {}
        for klass in klasses:
            categories = {
                (categoria.get("id")): categoria.get("nome")
                for categoria in (klass.get("categorias") or [])
            }
            classes_meta[klass.get("id")] = {
                "name": (klass.get("nome") or ""),
                "categories": categories,
            }

        for serie in (resultado.get("series") or []):
            value = None
            for raw_value in (serie.get("serie") or {}).values():
                value = _safe_parse_numeric(raw_value)
                if value is not None:
                    break
            if value is None:
                continue

            serie_classes = serie.get("classificacoes") or []
            # esperamos exatamente 1 classificação (sexo OU idade),
            # pois pedimos separadamente para evitar soma duplicada
            if len(serie_classes) != 1:
                # se vier mais de uma por alguma razão, pulamos para não enviesar
                continue

            sc = serie_classes[0]
            class_id = sc.get("id")
            category_id = sc.get("categoria")
            meta = classes_meta.get(class_id) or {}
            class_name = (meta.get("name") or "").lower()
            if target_class_name not in class_name:
                continue
            label = (meta.get("categories") or {}).get(category_id, str(category_id))
            out[label] = out.get(label, 0) + value
    return out


@lru_cache(maxsize=4096)
def fetch_demographics_by_code(code: str | None) -> Dict[str, Any] | None:
    if not code:
        return None

    local_entry = _local_lookup_by_code(code)
    payload_total = None
    payload_sex = None
    payload_age = None

    # 1) TOTAL (sem classificação) — preferível para obter 'total' diretamente
    try:
        url_total = _build_demographics_url(code, classificacoes=None)
        r_total = requests.get(url_total, timeout=15)
        r_total.raise_for_status()
        payload_total = r_total.json()
    except Exception as exc:
        _log("ibge.demographics_total_failed", code=code, error=str(exc))
        payload_total = None

    # 2) SEXO (apenas sexo; sem idade)
    try:
        url_sex = _build_demographics_url(code, classificacoes={SEX_CLASS_ID: SEX_IDS})
        r_sex = requests.get(url_sex, timeout=15)
        r_sex.raise_for_status()
        payload_sex = r_sex.json()
    except Exception as exc:
        _log("ibge.demographics_sex_failed", code=code, error=str(exc))
        payload_sex = None

    # 3) IDADE (apenas faixas etárias consolidadas; sem sexo)
    try:
        url_age = _build_demographics_url(code, classificacoes={AGE_CLASS_ID: AGE_BAND_IDS})
        r_age = requests.get(url_age, timeout=15)
        r_age.raise_for_status()
        payload_age = r_age.json()
    except Exception as exc:
        _log("ibge.demographics_age_failed", code=code, error=str(exc))
        payload_age = None

    # ---- Parse dos dados
    total = _parse_total_from_payload(payload_total)

    sex_breakdown: Dict[str, int] = _parse_breakdown(payload_sex, target_class_name="sexo") if payload_sex else {}
    age_breakdown: Dict[str, int] = _parse_breakdown(payload_age, target_class_name="idade") if payload_age else {}

    # fallback total (tabela legacy) se necessário
    if total is None:
        total = fetch_population_legacy(code)

    # Se nada deu certo:
    if total is None and not sex_breakdown and not age_breakdown:
        if local_entry and local_entry.get("population") is not None:
            return {
                "code": code,
                "total": local_entry["population"],
                "sex": {},
                "age": {},
                "period": local_entry.get("population_year"),
                "raw": None,
            }
        return None
    period_value = None
    if payload_total and isinstance(payload_total, list) and payload_total:
        first_result = payload_total[0].get("resultados", [])
        if first_result:
            period_keys = []
            for serie in first_result[0].get("series", []):
                period_keys.extend(list((serie.get("serie") or {}).keys()))
            if period_keys:
                period_value = period_keys[-1]
    if not period_value and local_entry:
        period_value = local_entry.get("population_year")
    return {
        "code": code,
        "total": total,
        "sex": sex_breakdown,
        "age": age_breakdown,
        "period": period_value,
        "raw": {
            "total": payload_total,
            "sex": payload_sex,
            "age": payload_age,
        },
    }


@lru_cache(maxsize=4096)
def fetch_population_legacy(code: str | None) -> int | None:
    if not code:
        return None
    try:
        resp = requests.get(IBGE_POPULATION_URL.format(code=code), timeout=10)
        resp.raise_for_status()
        payload = resp.json()
    except Exception as exc:
        _log("ibge.population_failed", code=code, error=str(exc))
        entry = _local_lookup_by_code(code)
        if entry:
            return entry.get("population")
        return None

    serie = (
        payload[0]
        .get("resultados", [{}])[0]
        .get("series", [{}])[0]
        .get("serie", {})
    )
    values = [
        _safe_parse_numeric(value)
        for value in serie.values()
        if value not in (None, "")
    ]
    values = [v for v in values if v is not None]
    if values:
        return values[-1]
    entry = _local_lookup_by_code(code)
    return entry.get("population") if entry else None


def fetch_demographics_by_city(city: str | None, state: str | None = None) -> Dict[str, Any] | None:
    code = resolve_municipality_code(city, state)
    if not code:
        entry = _local_lookup_by_city(city, state)
        if entry and entry.get("population") is not None:
            return {
                "code": entry["code"],
                "total": entry["population"],
                "sex": {},
                "age": {},
                "raw": None,
            }
        return None
    data = fetch_demographics_by_code(code)
    if data:
        return data
    total = fetch_population_legacy(code)
    if total is None:
        entry = _local_lookup_by_code(code)
        if entry and entry.get("population") is not None:
            return {
                "code": code,
                "total": entry["population"],
                "sex": {},
                "age": {},
                "raw": None,
            }
        return None
    return {"code": code, "total": total, "sex": {}, "age": {}, "raw": None}
